import os
import pandas as pd
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import openpyxl
import logging
from openpyxl.utils import get_column_letter

# 配置日志系统（如果单独运行此文件）
if __name__ == "__main__":
    # 创建基础日志配置
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler("vocab_export.log")
        ]
    )
    logger = logging.getLogger()

# 假设在其他模块中已经配置了logger
logger = logging.getLogger(__name__)

def export_vocab_to_excel(vocab_list, filename=None):
    """
    将词汇列表导出为美观、舒适的Excel文件
    
    参数:
        vocab_list: 词汇字典列表
        filename: 自定义文件名（可选）
    
    返回:
        str: 生成的文件路径
    """
    try:
        if not vocab_list:
            logger.warning("尝试导出空词汇列表，跳过导出")
            return None
        
        logger.info(f"开始导出词汇表到Excel，共 {len(vocab_list)} 个词汇项")
        
        # 创建数据框架
        df = pd.DataFrame(vocab_list)
        
        # 确保所有必需的列都存在
        required_columns = ['word', 'pos', 'definition', 'definition-ch', 'common-usage']
        for col in required_columns:
            if col not in df.columns:
                logger.warning(f"缺失列 '{col}'，创建空列")
                df[col] = ""
        
        # 添加额外列
        df["example"] = ""  # 留空用于用户添加例句
        df["mastery"] = ""  # 留空用于标记掌握程度
        df["notes"] = ""    # 留空用于添加笔记
        
        # 确保常见用法列格式化为以"|"分隔的字符串
        if 'common-usage' in df.columns:
            df['common-usage'] = df['common-usage'].apply(
                lambda x: ' | '.join(x) if isinstance(x, list) else str(x)
            )
        
        # 设置默认文件名
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"vocabulary_{timestamp}.xlsx"
        
        # 确保导出目录存在
        os.makedirs("exports", exist_ok=True)
        filepath = os.path.join("exports", filename)
        
        # 定义列顺序和标题
        column_order = [
            'word', 'pos', 'definition', 'definition-ch', 
            'common-usage', 'example', 'mastery', 'notes'
        ]
        
        # 重新排序列
        df = df[column_order]
        
        # 导出Excel
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Vocabulary', index=False)
            
            # 获取工作簿和工作表
            workbook = writer.book
            worksheet = writer.sheets['Vocabulary']
            
            # 设置列宽 - 优化美观性
            column_widths = {
                'A': 22,  # Word - 足够容纳大多数单词
                'B': 8,   # POS - 词性缩写
                'C': 42,  # Definition (English) - 中等宽度
                'D': 25,  # Definition (Chinese) - 中文通常更简洁
                'E': 55,  # Common Usage - 以"|"分隔的用法
                'F': 50,  # Example - 用户添加的例句
                'G': 12,  # Mastery - 掌握程度标记
                'H': 40   # Notes - 笔记
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # 设置标题行样式 - 专业而优雅
            header_fill = PatternFill(
                start_color='4F81BD',  # 专业蓝色
                end_color='4F81BD', 
                fill_type='solid'
            )
            header_font = Font(bold=True, color='FFFFFF', size=12)  # 白色加粗字体
            center_alignment = Alignment(horizontal='center', vertical='center')  # 水平垂直居中
            
            # 应用样式到标题行(第一行)
            for row in worksheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_alignment  # 标题行居中
                    # 设置标题行高度
                    worksheet.row_dimensions[1].height = 22
            
            # 设置数据行字体 - 更易读
            base_font = Font(name='Calibri', size=11)
            word_font = Font(name='Calibri', size=11, bold=True)  # 单词加粗
            
            # 定义优雅的边框
            thin_border = Border(
                left=Side(style='thin', color='D0D0D0'),
                right=Side(style='thin', color='D0D0D0'),
                top=Side(style='thin', color='D0D0D0'),
                bottom=Side(style='thin', color='D0D0D0')
            )
            
            # 设置行高和单元格样式 - 所有元素水平垂直居中
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), 2):
                # 设置行高（基础行高）
                worksheet.row_dimensions[row_idx].height = 24
                
                # 交替行颜色
                if row_idx % 2 == 0:
                    row_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # 白色
                else:
                    row_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')  # 浅灰色
                
                for cell in row:
                    # 应用填充和边框
                    cell.fill = row_fill
                    cell.border = thin_border
                    
                    # 设置字体
                    if cell.column_letter == 'A':  # 单词列
                        cell.font = word_font
                    else:
                        cell.font = base_font
                    
                    # 设置所有单元格水平垂直居中
                    cell.alignment = center_alignment
            
            # 冻结标题行
            worksheet.freeze_panes = 'A2'
            
            # 添加筛选器
            worksheet.auto_filter.ref = f"A1:H1"
            
            # 添加文件元数据
            workbook.properties.title = "词汇表"
            workbook.properties.subject = "从文章中提取的词汇"
            workbook.properties.author = "词汇提取工具"
            workbook.properties.creator = "词汇提取工具"
            
            # 设置工作表标签颜色
            worksheet.sheet_properties.tabColor = "4F81BD"
        
        logger.info(f"成功导出美观的词汇表: {filepath}")
        return filepath
        
    except PermissionError:
        logger.error("文件访问权限错误，请确保文件未被其他程序打开")
        return None
    except Exception as e:
        logger.error(f"导出Excel时出错: {str(e)}", exc_info=True)
        return None


def export_to_csv(vocab_list, filename=None):
    """
    将词汇列表导出为CSV文件（备选格式）
    
    参数:
        vocab_list: 词汇字典列表
        filename: 自定义文件名（可选）
    
    返回:
        str: 生成的文件路径
    """
    try:
        if not vocab_list:
            logger.warning("尝试导出空词汇列表到CSV，跳过导出")
            return None
        
        logger.info(f"开始导出词汇表到CSV，共 {len(vocab_list)} 个词汇项")
        
        # 创建数据框架
        df = pd.DataFrame(vocab_list)
        
        # 确保常见用法列格式化为以"|"分隔的字符串
        if 'common-usage' in df.columns:
            df['common-usage'] = df['common-usage'].apply(
                lambda x: ' | '.join(x) if isinstance(x, list) else str(x)
            )
        
        # 设置默认文件名
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"vocabulary_{timestamp}.csv"
        
        # 确保导出目录存在
        os.makedirs("exports", exist_ok=True)
        filepath = os.path.join("exports", filename)
        
        # 导出CSV
        df.to_csv(filepath, index=False, encoding='utf-8-sig')
        
        logger.info(f"成功导出CSV词汇表: {filepath}")
        return filepath
        
    except Exception as e:
        logger.error(f"导出CSV时出错: {str(e)}", exc_info=True)
        return None


# 在main函数中添加测试代码
if __name__ == "__main__":
    # 示例词汇数据
    sample_vocab = [
        {
            "word": "aesthetic",
            "pos": ".adj",
            "definition": "concerned with beauty or the appreciation of beauty",
            "definition-ch": "美学的；审美的",
            "common-usage": ["aesthetic appeal", "design aesthetic"]
        },
        {
            "word": "eloquent",
            "pos": ".adj",
            "definition": "fluent or persuasive in speaking or writing",
            "definition-ch": "雄辩的；有口才的",
            "common-usage": ["an eloquent speaker", "eloquent expression of ideas"]
        },
        {
            "word": "paradigm",
            "pos": ".n",
            "definition": "a typical example or pattern of something; a model",
            "definition-ch": "范例；典范",
            "common-usage": ["scientific paradigm", "shift in paradigm"]
        },
        {
            "word": "ubiquitous",
            "pos": ".adj",
            "definition": "present, appearing, or found everywhere",
            "definition-ch": "无所不在的；普遍存在的",
            "common-usage": ["ubiquitous technology", "ubiquitous presence"]
        },
        {
            "word": "ephemeral",
            "pos": ".adj",
            "definition": "lasting for a very short time",
            "definition-ch": "短暂的；瞬息的",
            "common-usage": ["ephemeral beauty", "ephemeral nature of fame"]
        }
    ]
    
    # 测试导出功能
    excel_file = export_vocab_to_excel(sample_vocab)
    if excel_file:
        print(f"美观的Excel词汇表已创建: {excel_file}")
    
    # 测试CSV导出功能
    csv_file = export_to_csv(sample_vocab)
    if csv_file:
        print(f"CSV词汇表已创建: {csv_file}")
    
    # 测试空列表处理
    empty_result = export_vocab_to_excel([])
    if not empty_result:
        print("空列表处理测试成功")